from openpyxl import load_workbook, formatting, styles
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import pandas as pd
import re
import xlrd
import argparse

#讀入相關的參數
ap = argparse.ArgumentParser()
ap.add_argument("-M", "--main", required=True, help="Path to the main Intel file.")
ap.add_argument("-S", "--sub", required=True, help="Path to the ball name file.")
args = vars(ap.parse_args()) 

#從第一份文件找出要的GPIO
def step1_sortGPIO(filePath, sheetName, GPIOCol):
    wb = load_workbook(filePath)
    ws = wb[sheetName]
    GPIOList0 = []
    for i in ws[GPIOCol]:
        GPIOList0.append(i.value)
    return GPIOList0

#從第二份文件找出要的GPIO 的 ball name
def step2_sortBallName(filePath, sheetName):

    data = pd.read_excel(filePath, sheet_name = sheetName).iloc[:,:2]

    GPIOList2 = []
    for idx in data.index:

        eachGPIO ={"GPIO": data.iloc[idx,1].split('/')[0].replace(" ", "") ,"ball" : data.iloc[idx,0]}
        GPIOList2.append(eachGPIO)
    return GPIOList2

#把第二份的ballname map過去第一份並產生dictionary
def step3_map_GPIO_ballName(GPIOList1, GPIOList2):
    ballNameCol = []
    for GPIO in GPIOList1:
        data ={"GPIO" : GPIO, "BallName" : ""}
        if type(GPIO) == str: #如果有list1有GPIO不是nan
            for i in GPIOList2:#跟GPIO
                if GPIO == i["GPIO"]:
                    data["BallName"] = i["ball"]
            ballNameCol.append(data)
        else:        
            ballNameCol.append(data)        
    return ballNameCol

#把整理好的GPIO 跟 ball name 放進Intel file李
def step4_insert_GPIO_ballName(mainFilePath, sheetName, ballNameCol):
    wb = load_workbook(mainFilePath)
    ws = wb[sheetName]
    ws.insert_cols(3,2) #從第三航插入兩列
    
    #開始insert
    num = 0
    
    #styles
    textFont = Font(name='Verdana', size=10)
    textAlignment = Alignment(vertical='center')
    
    titlefont = Font(name='Verdana', size=10, color = "FF00B050", family=2.0, bold=True)
    titlefontAlignment = Alignment(horizontal='center',vertical='center')
    
    
    for i in ws['C']:
        i.value = ballNameCol[num]['GPIO']
        i.font = textFont
        i.alignment = textAlignment
        
        ws.cell(row= num +1 , column= 4).value = ballNameCol[num]['BallName'] #保持num和i是同一列
        ws.cell(row= num +1 , column= 4).font = textFont
        ws.cell(row= num +1 , column= 4).alignment = textAlignment
        
        num += 1
        
    ws['C2'] = "GPIO\n(reference)"
    ws['D2'] = "Ball Name"
    ws['D2'].font = titlefont
    ws['D2'].alignment = titlefontAlignment
    ws['C2'].font = titlefont
    ws['C2'].alignment = titlefontAlignment
    
    wb.save(f"{mainFilePath}_+ballName.xlsx")


#step1
#先找出要的GPIO
# mainFilePath, sheetName, GPIOCol = '627075_ADL_P_PCH_GPIO_IS_Rev1p0.xlsx', 'GPIO Implementation', "B"
mainFilePath, sheetName, GPIOCol = args["main"], 'GPIO Implementation', "B"
try:
    GPIOList1 = step1_sortGPIO(mainFilePath, sheetName, GPIOCol)
    print("step 1: Extract GPIO List Success.")
except:
    print("Clear Data Fail, check step 1: 'function step11_sortGPIO' ")
    
#step2
subfilePath, sheetName = args["sub"], 'Pinlist'
try:
    GPIOList2 = step2_sortBallName(subfilePath, sheetName)
    print("step 2: Extract CPU Ball Name Success.")
except:
    print("Clear Data Fail, check step 2: 'step2_sortBallName' ")

    
#step3
try:
    ballNameCol = step3_map_GPIO_ballName(GPIOList1, GPIOList2)
    print("step 3: Mapping Ballname Success.")
except:
    print("Mapping Ballname Fail, check step 3: 'step3_map_GPIO_ballName' ")
    
#step4 insert ballname

mainFilePath, sheetName = args["main"], 'GPIO Implementation'

try:
    step4_insert_GPIO_ballName(mainFilePath, sheetName, ballNameCol)
    print("step 4: Insert Ballname Success.")
except:
    print("Inserting Ballname Fail, check step 4: 'step4_insert_GPIO_ballName' ")
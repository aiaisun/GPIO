import re
import pandas as pd
import argparse
from openpyxl import load_workbook, formatting, styles
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

#讀入相關的參數
ap = argparse.ArgumentParser()
ap.add_argument("-t", "--txt", required=True, help="Path to the txt of the TXT file.")
args = vars(ap.parse_args())


#step1: 開啟project TXT
def step1_openProjectTXT(txtFilePath):    
    f = open(txtFilePath)
    return f

def step2_sortProjectTXT(f):
    rawData = []
    for i in f:
        row = i.replace("\n","")#去掉換行符號
        rawData.append(row)
    
    #找出個$NETS並清除不需要的資料
    keyIDX = []
    for i in rawData:
        if re.findall(f"\$(NETS|END)", i , re.I):
    #         print(re.findall(f"\$.*", i, re.I))
            keyIDX.append(rawData.index(i))
    rawData = rawData[keyIDX[0] +1: keyIDX[1]]
    
    return rawData


# 把TXT整理成dictionary
def step3_sortTXTtoDic(rawData):
    netName, netData = "", ""
    netInfo = {}
    clearData = {}
    
    # 處理txt檔 分好net name 跟 GPIO資料
    for line in rawData:
        #有net name的那一行
        if re.findall(f".+\;", line): 
            
            netName = re.findall(f".+\;", line)[0].replace(";", "")#分號前是net name

            #處理 net data 
            netData = re.findall(f"\;.+", line)[0].replace(";", "")#分號後是net data
            netData = netData.split()
            netData = ",".join(netData)

            netInfo = { netName : netData }
            clearData.update(netInfo)
        #沒有net name的那一行，留在同一個net name  
        else:
            line = ",".join(line.split())#用空格白切割 再用,分割

            if clearData.__contains__(netName):
                clearData[netName] += line
            else:
                print("erroe", line)
    return clearData


#從第一份文件找出要的ball name
def step4_extractIntelDoc(mainFilePath, sheetName):
    wb = load_workbook(mainFilePath)
    ws = wb[sheetName]
    ws.insert_cols(5,1)
    return wb, ws

#mapping project and Intel
def step5_mapping_netname(clearData, ws):
    textFont = Font(name='Verdana', size=10)
    textAlignment = Alignment(vertical='center')

    for netName, netData in clearData.items():

        for i in ws['D']:#intel裡面的ball name

            if type(i.value) == str:
                result = re.findall(f"(UCPU1\.{i.value}\,|UCPU1\.{i.value}$)", netData, re.I)

                if result:
                    #確認有無重複
                    if not bool(ws.cell(row = row , column= 5).value):
                        row = i.row #i.coordinate
                        ws.cell(row = row , column= 5).value = netName
                        ws.cell(row = row , column= 5).font = textFont
                        ws.cell(row = row , column= 5).alignment = textAlignment
                    else:
                        ws.insert_cols(5,1)
                        ws.cell(row = row , column= 6).value = netName
                        ws.cell(row = row , column= 6).font = textFont
                        ws.cell(row = row , column= 6).alignment = textAlignment        
                        

            else:
                row = i.row
                ws.cell(row= row , column= 5).value = ""

    titlefont = Font(name='Verdana', size=10, color = "FF00B050", family=2.0, bold=True)
    titlefontAlignment = Alignment(horizontal='center',vertical='center', wrapText=True) 
    titleborder = Border(left   = styles.Side(border_style='medium', color='FF000000'),
                         right  = styles.Side(border_style='medium', color='FF000000'),
                         top    = styles.Side(border_style='medium', color='FF000000'),
                         bottom = styles.Side(border_style='medium', color='FF000000'),
                         )
    ws['E2'] = "Project"
    ws['E2'].font = titlefont
    ws['E2'].alignment = titlefontAlignment
    ws['E2'].border = titleborder
    
#存檔
def step6_saveFile(txtFilePath, wb):
    txtFilePath = txtFilePath.replace('.txt', "")
    wb.save(f"{txtFilePath}.xlsx")
    return txtFilePath

filepath = args["txt"]
print(filepath)
f = open(filepath)
print(f)

# g = open("ORB_ADL_LPDDR5_1125b.txt")
# print(g)

# try:
#     f = step1_openProjectTXT("ORB_ADL_LPDDR5_1125b.txt")
#     print(f"Success: Step1: Open Project TXT- {filepath} -DONE")
# except:
#     print(f"Fail: check step 1: 'function step1_sortProjectTXT'")

# try:
#     rawData = step2_sortProjectTXT(f)
#     print(f"Success: Step2: Sort Project TXT -DONE")
# except:
#     print(f"Fail: check step 2: 'function step2_sortProjectTXT'")
    
# #step3: 把TXT整理成dictionary  
# try:
#     clearData = step3_sortTXTtoDic(rawData)
#     print(f"Success: Step3: Sort Project TXT to Dic-DONE")
# except:
#     print(f"Fail: check step 3: 'function step3_sortTXTtoDic'")
    
# #step4
# #先找出要的Ball Name
# mainFilePath, sheetName = 'Intel_ballname.xlsx', 'GPIO Implementation'
# try:
#     wb, ws = step4_extractIntelDoc(mainFilePath, sheetName)
#     print(f"Success: Step4: Extract Intel Doc -DONE")
# except:
#     print(f"Fail: check step 4: 'function step4_extractIntelDoc'")
    
# #step5
# #mapping Project and Intel
# try:
#     step5_mapping_netname(clearData, ws)
#     print(f"Success: Step5: Mapping Project Net Name -DONE")
# except:
#     print(f"Fail: check step 5: 'function step5_mapping_netname'")

# #step6
# #存檔   
# try:
#     txtFilePath = step6_saveFile(filepath, wb)
#     print(f"Success: Step6: Mapped File {txtFilePath} Saved -DONE")
# except:
#     print(f"Fail: check step 6: 'function step6_saveFile'")